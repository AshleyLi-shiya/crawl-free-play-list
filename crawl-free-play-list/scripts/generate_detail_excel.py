#!/usr/bin/env python3
"""
Generate detailed Excel with all free content from 4 platforms.
Fields: 平台, 类型, 剧名, 剧目年份, 演员, 地区, 集数, 题材, 男女频
"""
import json
import re
import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(SCRIPT_DIR)
DATA_DIR = os.path.join(PROJECT_ROOT, 'data')
OUTPUT_DIR = os.path.join(PROJECT_ROOT, 'output')
os.makedirs(DATA_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)

# ========== Gender Classification Logic ==========

MALE_GENRES = {'军旅', '谍战', '武侠', '玄幻', '仙侠', '抗战', '犯罪', '警匪',
               '动作', '商战', '主旋律', '枪战', '灾难', '竞技', '神话', '魔幻',
               '战争', '罪案', '刑侦', '军事'}

FEMALE_GENRES = {'言情', '宫廷', '甜宠', '宅斗', '宫斗', '纯爱'}

FEMALE_LEANING_GENRES = {'爱情', '穿越', '偶像', '浪漫', '婚姻'}

MALE_LEANING_GENRES = {'历史', '科幻', '悬疑', '探险', '冒险'}

FEMALE_KEYWORDS = [
    '甜宠', '霸总', '总裁', '闪婚', '离婚', '千金', '灰姑娘', '公主',
    '后宫', '嫡女', '宠妻', '甜蜜', '暖男', '傲娇', '宠爱', '婆媳',
    '真千金', '替嫁', '冲喜', '萌宝', '甜妻', '豪门', '腹黑', '契约婚',
    '宫斗', '宅斗', '小姐', '夫人', '王妃', '皇后', '妃子', '侧妃',
    '太子妃', '世子妃', '王后', '贵妃', '心尖宠', '团宠', '娘娘',
    '重生之', '穿越之', '嫁给', '和离', '休妻', '下堂', '冷面', '冰山',
    '权臣', '相府', '侯府', '将军府', '丞相', '庶女', '嫡妹',
    '闺蜜', '恋爱', '结婚', '婚后', '新婚', '蜜恋', '暗恋',
    '初恋', '虐恋', '甜文', '宠文', '追妻', '情深', '倾城',
    '美人', '佳人', '红颜', '倾国', '小娇妻', '娇妻',
    '姐弟恋', '先婚后爱', '日久生情', '双向奔赴',
    '闪婚老公', '隐婚', '逼婚', '财阀', 'CEO',
    '男神', '校草', '学长', '竹马', '青梅',
]

MALE_KEYWORDS = [
    '兄弟', '热血', '江湖', '修仙', '王者', '称霸', '修炼', '神功',
    '大佬', '黑帮', '特种兵', '功夫', '武功', '枭雄', '龙王', '战神',
    '至尊', '霸主', '天尊', '魔尊', '帝尊', '剑仙', '道长',
    '赘婿', '上门女婿', '退婚', '废物', '逆天', '崛起', '觉醒',
    '无敌', '最强', '巅峰', '吊打', '碾压', '秒杀', '横扫',
    '兵王', '佣兵', '雇佣兵', '杀手', '刺客', '狙击手',
    '首富', '商业帝国', '权势', '权力', '复仇', '重生都市',
    '全能', '系统', '金手指', '外挂', '神豪', '神医',
    '战士', '勇士', '猛将', '战将', '军师',
    '黑道', '教父', '帮主', '老大', '堂主',
    '太子', '皇子', '世子', '侯爷', '将军',
    '龙', '虎', '狼', '鹰',
]


def classify_gender(title, genres_list, description=''):
    """Classify content as 男频/女频/中性. Returns one of 3 labels."""
    text = title + ' ' + description
    male_score = 0
    female_score = 0

    for g in genres_list:
        g = g.strip()
        if g in MALE_GENRES:
            male_score += 3
        elif g in FEMALE_GENRES:
            female_score += 3
        elif g in MALE_LEANING_GENRES:
            male_score += 1.5
        elif g in FEMALE_LEANING_GENRES:
            female_score += 1.5

    for kw in FEMALE_KEYWORDS:
        if kw in title:
            female_score += 2.5
        elif kw in text:
            female_score += 1

    for kw in MALE_KEYWORDS:
        if kw in title:
            male_score += 2.5
        elif kw in text:
            male_score += 1

    if re.search(r'(她|女|娘|姐|妹|妻|嫂|媳|母|婆|后|妃|姑|姨|丫鬟|公主|皇后|王妃|太后|贵妃)', title):
        female_score += 1
    if re.search(r'(他|男|爷|哥|弟|夫|父|叔|伯|帝|王|侯|将|兵|士|皇帝|大人)', title):
        male_score += 1

    # Simplify to 3 categories as user requested
    if male_score > female_score and male_score >= 1.5:
        return '男频'
    elif female_score > male_score and female_score >= 1.5:
        return '女频'
    else:
        return '中性'


# ========== Data Loading ==========

def load_mgtv():
    """Load Mango TV data (regular TV dramas + movies)"""
    filepath = os.path.join(DATA_DIR, 'mgtv_data.json')
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            data = json.load(f)
    except FileNotFoundError:
        print(f"  [WARN] {filepath} not found, skipping MangoTV")
        return []

    rows = []
    # TV Dramas (regular only, short dramas excluded)
    for item in data['tvDramas']['regular']:
        genres_list = item.get('kind', [])
        genres_str = ','.join(genres_list)
        title = item.get('title', '')
        story = item.get('story', '')
        label = classify_gender(title, genres_list, story)
        rows.append({
            '平台': '芒果TV',
            '类型': '电视剧',
            '剧名': title,
            '剧目年份': item.get('year', ''),
            '演员': item.get('subtitle', ''),
            '地区': '',  # Mango TV data doesn't have region
            '集数': item.get('updateInfo', ''),
            '题材': genres_str,
            '男女频': label,
        })

    # Movies
    for item in data['movies']:
        genres_list = item.get('kind', [])
        genres_str = ','.join(genres_list)
        title = item.get('title', '')
        story = item.get('story', '')
        label = classify_gender(title, genres_list, story)
        rows.append({
            '平台': '芒果TV',
            '类型': '电影',
            '剧名': title,
            '剧目年份': item.get('year', ''),
            '演员': item.get('subtitle', ''),
            '地区': '',
            '集数': item.get('updateInfo', ''),
            '题材': genres_str,
            '男女频': label,
        })

    return rows


def load_iqiyi():
    """Load iQIYI data (full details)"""
    filepath = os.path.join(DATA_DIR, 'iqiyi_full_details.json')
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            data = json.load(f)
    except FileNotFoundError:
        print(f"  [WARN] {filepath} not found, skipping iQIYI")
        return []

    rows = []
    for category, type_label in [('tvDramas', '电视剧'), ('movies', '电影')]:
        for item in data.get(category, []):
            title = item.get('title', '')
            genre_str = item.get('genre', '')
            genres_list = [g.strip() for g in genre_str.split(',') if g.strip()]
            desc = item.get('description', '')
            label = classify_gender(title, genres_list, desc)
            rows.append({
                '平台': '爱奇艺',
                '类型': type_label,
                '剧名': title,
                '剧目年份': item.get('year', ''),
                '演员': item.get('actors', ''),
                '地区': item.get('region', ''),
                '集数': item.get('episodes', ''),
                '题材': genre_str,
                '男女频': label,
            })

    return rows


def load_tencent():
    """Load Tencent Video data"""
    filepath = os.path.join(DATA_DIR, 'tencent_details.json')
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            data = json.load(f)
    except FileNotFoundError:
        print(f"  [WARN] {filepath} not found, skipping Tencent Video")
        return []

    rows = []
    for category, type_label in [('tvDramas', '电视剧'), ('movies', '电影')]:
        for item in data.get(category, []):
            title = item.get('title', '')
            genre_str = item.get('genre', '')
            genres_list = [g.strip() for g in genre_str.split(',') if g.strip()]
            desc = item.get('desc', '')
            # Tencent actors are space-separated, normalize to comma
            actors_raw = item.get('actors', '')
            actors = actors_raw.replace(' ', ',') if actors_raw else ''
            label = classify_gender(title, genres_list, desc)
            rows.append({
                '平台': '腾讯视频',
                '类型': type_label,
                '剧名': title,
                '剧目年份': item.get('year', ''),
                '演员': actors,
                '地区': item.get('area', ''),
                '集数': item.get('episodes', ''),
                '题材': genre_str,
                '男女频': label,
            })

    return rows


def load_youku():
    """Load Youku data (enriched details)"""
    filepath = os.path.join(DATA_DIR, 'youku_details.json')
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            data = json.load(f)
    except FileNotFoundError:
        print(f"  [WARN] {filepath} not found, skipping Youku")
        return []

    rows = []
    for category, type_label in [('tvDramas', '电视剧'), ('movies', '电影')]:
        for item in data.get(category, []):
            title = item.get('title', '')
            genre_str = item.get('genre', '')
            genres_list = [g.strip() for g in genre_str.split(',') if g.strip()]
            desc = item.get('desc', '')
            label = classify_gender(title, genres_list, desc)
            rows.append({
                '平台': '优酷',
                '类型': type_label,
                '剧名': title,
                '剧目年份': item.get('year', ''),
                '演员': item.get('actors', ''),
                '地区': item.get('region', ''),
                '集数': item.get('episodes', ''),
                '题材': genre_str,
                '男女频': label,
            })

    return rows


# ========== Excel Generation ==========

def generate_excel(all_rows, output_path):
    wb = openpyxl.Workbook()

    # Styles
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_font = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Microsoft YaHei", size=10)
    male_fill = PatternFill(start_color="DAEEF3", end_color="DAEEF3", fill_type="solid")
    female_fill = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")
    neutral_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="B0B0B0"),
        right=Side(style="thin", color="B0B0B0"),
        top=Side(style="thin", color="B0B0B0"),
        bottom=Side(style="thin", color="B0B0B0"),
    )

    fields = ['平台', '类型', '剧名', '剧目年份', '演员', '地区', '集数', '题材', '男女频']
    col_widths = [12, 8, 28, 10, 40, 10, 14, 24, 10]
    # columns that should be center-aligned
    center_cols = {0, 1, 3, 5, 6, 8}  # 平台,类型,年份,地区,集数,男女频

    # Gender fill mapping
    gender_fills = {'男频': male_fill, '女频': female_fill, '中性': neutral_fill}

    # --- Sheet 1: All data ---
    ws = wb.active
    ws.title = "全部详情"

    # Header row
    for i, h in enumerate(fields):
        cell = ws.cell(row=1, column=i + 1, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Data rows
    for r_idx, row in enumerate(all_rows, start=2):
        for c_idx, field in enumerate(fields):
            val = row.get(field, '')
            cell = ws.cell(row=r_idx, column=c_idx + 1, value=val)
            cell.font = data_font
            cell.alignment = center_align if c_idx in center_cols else left_align
            cell.border = thin_border
            # Color the 男女频 column
            if c_idx == 8:
                fill = gender_fills.get(val)
                if fill:
                    cell.fill = fill

    # Column widths
    for i, w in enumerate(col_widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = w

    # Freeze header
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(fields))}{len(all_rows) + 1}"

    # --- Sheet 2: Summary stats ---
    ws2 = wb.create_sheet("汇总统计")
    title_font = Font(name="Microsoft YaHei", size=14, bold=True, color="2F5496")
    subtitle_font = Font(name="Microsoft YaHei", size=11, bold=True, color="333333")

    ws2.merge_cells("A1:G1")
    c = ws2.cell(row=1, column=1, value="四平台免费内容汇总统计")
    c.font = title_font
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 35

    # Count stats
    platforms = ['芒果TV', '爱奇艺', '腾讯视频', '优酷']
    types = ['电视剧', '电影']

    # Build stats dict
    stats = {}
    for p in platforms:
        stats[p] = {}
        for t in types:
            subset = [r for r in all_rows if r['平台'] == p and r['类型'] == t]
            male_c = sum(1 for r in subset if r['男女频'] == '男频')
            female_c = sum(1 for r in subset if r['男女频'] == '女频')
            neutral_c = sum(1 for r in subset if r['男女频'] == '中性')
            stats[p][t] = {
                'total': len(subset),
                '男频': male_c,
                '女频': female_c,
                '中性': neutral_c,
            }

    # TV Drama summary table
    ws2.merge_cells("A3:G3")
    ws2.cell(row=3, column=1, value="一、免费电视剧").font = subtitle_font

    sum_headers = ["平台", "免费总数", "男频", "男频占比", "女频", "女频占比", "中性"]
    for i, h in enumerate(sum_headers):
        cell = ws2.cell(row=4, column=i + 1, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    for r_idx, p in enumerate(platforms, start=5):
        s = stats[p]['电视剧']
        total = s['total']
        m_pct = f"{s['男频']/total*100:.1f}%" if total else "0%"
        f_pct = f"{s['女频']/total*100:.1f}%" if total else "0%"
        vals = [p, total, s['男频'], m_pct, s['女频'], f_pct, s['中性']]
        fills_row = [None, None, male_fill, male_fill, female_fill, female_fill, neutral_fill]
        for c_idx, v in enumerate(vals):
            cell = ws2.cell(row=r_idx, column=c_idx + 1, value=v)
            cell.font = data_font
            cell.alignment = center_align
            cell.border = thin_border
            if fills_row[c_idx]:
                cell.fill = fills_row[c_idx]

    # Movie summary table
    movie_start = 5 + len(platforms) + 1
    ws2.merge_cells(f"A{movie_start}:G{movie_start}")
    ws2.cell(row=movie_start, column=1, value="二、免费电影").font = subtitle_font

    for i, h in enumerate(sum_headers):
        cell = ws2.cell(row=movie_start + 1, column=i + 1, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    for r_idx, p in enumerate(platforms, start=movie_start + 2):
        s = stats[p]['电影']
        total = s['total']
        m_pct = f"{s['男频']/total*100:.1f}%" if total else "0%"
        f_pct = f"{s['女频']/total*100:.1f}%" if total else "0%"
        vals = [p, total, s['男频'], m_pct, s['女频'], f_pct, s['中性']]
        fills_row = [None, None, male_fill, male_fill, female_fill, female_fill, neutral_fill]
        for c_idx, v in enumerate(vals):
            cell = ws2.cell(row=r_idx, column=c_idx + 1, value=v)
            cell.font = data_font
            cell.alignment = center_align
            cell.border = thin_border
            if fills_row[c_idx]:
                cell.fill = fills_row[c_idx]

    # Grand total
    grand_start = movie_start + 2 + len(platforms) + 1
    ws2.merge_cells(f"A{grand_start}:G{grand_start}")
    ws2.cell(row=grand_start, column=1, value="三、汇总").font = subtitle_font

    total_headers = ["平台", "电视剧", "电影", "合计"]
    for i, h in enumerate(total_headers):
        cell = ws2.cell(row=grand_start + 1, column=i + 1, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    for r_idx, p in enumerate(platforms, start=grand_start + 2):
        tv_n = stats[p]['电视剧']['total']
        mv_n = stats[p]['电影']['total']
        vals = [p, tv_n, mv_n, tv_n + mv_n]
        for c_idx, v in enumerate(vals):
            cell = ws2.cell(row=r_idx, column=c_idx + 1, value=v)
            cell.font = data_font
            cell.alignment = center_align
            cell.border = thin_border

    # Total row
    total_row = grand_start + 2 + len(platforms)
    all_tv = sum(stats[p]['电视剧']['total'] for p in platforms)
    all_mv = sum(stats[p]['电影']['total'] for p in platforms)
    vals = ['合计', all_tv, all_mv, all_tv + all_mv]
    bold_font = Font(name="Microsoft YaHei", size=10, bold=True)
    for c_idx, v in enumerate(vals):
        cell = ws2.cell(row=total_row, column=c_idx + 1, value=v)
        cell.font = bold_font
        cell.alignment = center_align
        cell.border = thin_border

    for i, w in enumerate([14, 12, 12, 12, 12, 12, 12], start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    wb.save(output_path)
    return stats


# ========== Main ==========

def main():
    print("Loading Mango TV data...")
    mgtv_rows = load_mgtv()
    print(f"  芒果TV: {sum(1 for r in mgtv_rows if r['类型']=='电视剧')} TV + {sum(1 for r in mgtv_rows if r['类型']=='电影')} movies")

    print("Loading iQIYI data...")
    iqiyi_rows = load_iqiyi()
    print(f"  爱奇艺: {sum(1 for r in iqiyi_rows if r['类型']=='电视剧')} TV + {sum(1 for r in iqiyi_rows if r['类型']=='电影')} movies")

    print("Loading Tencent data...")
    tencent_rows = load_tencent()
    print(f"  腾讯视频: {sum(1 for r in tencent_rows if r['类型']=='电视剧')} TV + {sum(1 for r in tencent_rows if r['类型']=='电影')} movies")

    print("Loading Youku data...")
    youku_rows = load_youku()
    print(f"  优酷: {sum(1 for r in youku_rows if r['类型']=='电视剧')} TV + {sum(1 for r in youku_rows if r['类型']=='电影')} movies")

    all_rows = mgtv_rows + iqiyi_rows + tencent_rows + youku_rows
    print(f"\nTotal: {len(all_rows)} items")

    output_path = os.path.join(OUTPUT_DIR, f"四平台免费内容详情-{datetime.now().strftime('%Y%m%d')}.xlsx")
    print(f"Generating Excel: {output_path}")
    stats = generate_excel(all_rows, output_path)

    # Print summary
    print("\n" + "=" * 60)
    print("汇总统计")
    print("=" * 60)
    for p in ['芒果TV', '爱奇艺', '腾讯视频', '优酷']:
        for t in ['电视剧', '电影']:
            s = stats[p][t]
            total = s['total']
            if total:
                print(f"  {p} {t}: {total}部 | 男频:{s['男频']}({s['男频']/total*100:.1f}%) 女频:{s['女频']}({s['女频']/total*100:.1f}%) 中性:{s['中性']}({s['中性']/total*100:.1f}%)")

    print(f"\nExcel saved: {output_path}")


if __name__ == '__main__':
    main()
